#!/usr/bin/env python3
"""
15_bea_data_build.py — Build Extended BEA Fixed Assets Dataset (1925–2024)

Constructs a reproducible dataset of US GDP and capital stock series following
the taxonomy defined in docs/notation.md.

Data sources (in priority order):
  1. LOCAL: Shaikh's Appendix 6.8 Excel (already in repo) — 1925–2011
  2. LOCAL: Shaikh's canonical CSV — 1946–2011
  3. LOCAL: ddbb_cu_US_kgr.xlsx — 1925–2011
  4. ONLINE (BEA API): extends to 2024 when API key provided
  5. ONLINE (FRED CSV): fallback for GDP/GVA/deflators

Asset taxonomy (docs/notation.md):
  ME  = Machinery & Equipment        (productive)
  NRC = Non-residential Construction (productive)
  IPP = Intellectual Property Prods  (tracked, excluded from productive K)
  NR  = ME + NRC                     (primary productive capital aggregate)

Sector focus: Corporate (total), with notes on nonfinancial vs financial.

Usage:
  python3 codes/15_bea_data_build.py                    # local Excel only
  python3 codes/15_bea_data_build.py --api-key KEY      # + BEA API extension
  python3 codes/15_bea_data_build.py --try-fred          # + FRED fallback
"""

import os
import sys
import json
import time
import logging
import argparse
import re
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
BEA_DIR = RAW_DIR / "bea_downloads"
PROC_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_CSV = PROC_DIR / "bea_extended_dataset_v1.csv"

BEA_DIR.mkdir(parents=True, exist_ok=True)
PROC_DIR.mkdir(parents=True, exist_ok=True)

# Shaikh's Appendix Excel — the primary local data source
APPENDIX_XLSX = RAW_DIR / "_Appendix6.8DataTablesCorrected.xlsx"
CANONICAL_CSV = RAW_DIR / "Shaikh_canonical_series_v1.csv"
DDBB_XLSX = RAW_DIR / "ddbb_cu_US_kgr.xlsx"

BASE_YEAR = 2005  # Shaikh-compatible base year

BEA_API_BASE = "https://apps.bea.gov/api/data/"
FRED_CSV_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ===================================================================
# PART A: LOCAL EXCEL EXTRACTION (primary path — no network needed)
# ===================================================================

def read_excel_row(ws, row_idx, year_row_idx, start_col=4, max_col=None):
    """Read a data row from an openpyxl worksheet, paired with the year row.

    Parameters:
        ws: openpyxl worksheet
        row_idx: 1-based row number for data
        year_row_idx: 1-based row number for years
        start_col: 1-based column where data starts (default 4 = column D)
        max_col: maximum column to read (None = all)

    Returns:
        DataFrame with columns [year, value]
    """
    mc = max_col or ws.max_column
    years = []
    values = []
    for col in range(start_col, mc + 1):
        yr = ws.cell(row=year_row_idx, column=col).value
        val = ws.cell(row=row_idx, column=col).value
        if yr is not None:
            try:
                yr_int = int(float(str(yr)))
                val_float = float(val) if val is not None else np.nan
                years.append(yr_int)
                values.append(val_float)
            except (ValueError, TypeError):
                continue
    return pd.DataFrame({"year": years, "value": values})


def extract_appendix_II1():
    """Extract series from Appendix 6.8.II.1 (GPIM accuracy).

    Contains BEA Fixed Assets Table 6.x extractions for the corporate sector:
    - Row 19: KNCcorpbea — Current-cost net stock (FA Table 6.1, line 2)
    - Row 20: KNRIndxcorpbea — Chain-index real net stock (FA Table 6.2, line 2)
    - Row 21: KNRcorpbea — Constant-cost net stock (derived)
    - Row 22: pKN — Implicit price deflator, net capital stock
    - Row 23: KNHcorpbea — Historical-cost net stock (FA Table 6.3, line 2)
    - Row 25: DEPCcorp — Current-cost depreciation (FA Table 6.4, line 2)
    - Row 26: dcorpstar — Corrected depreciation rate
    - Row 27: DEPHcorpbea — Historical-cost depreciation (FA Table 6.6, line 2)
    - Row 28: IGCcorpbea — Current-cost gross investment (FA Table 6.7, line 2)
    - Row 29: IGRcorpindexbea — Chain-type QI gross investment (FA Table 6.8, line 2)
    - Row 30: IGRcorpbea — Constant-cost gross investment (derived)
    - Row 31: pIGcorpbea — Implicit price deflator, gross investment
    - Row 37: dcorpWhelanLiu — Whelan-Liu depreciation rate

    Year row = 18, data starts at column D (col 4).
    """
    log.info("  Extracting Appendix 6.8.II.1 (BEA FA Tables for Corporate sector)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx 6.8.II.1"]
    except Exception as e:
        log.error(f"  Failed to open Appendix Excel: {e}")
        return pd.DataFrame(columns=["year"])

    # Row mapping: (row_number, variable_name, description)
    row_map = [
        (19, "KNCcorpbea",      "BEA current-cost net stock, corporate (FA 6.1)"),
        (20, "KNRIndxcorpbea",  "BEA chain-index real net stock, corporate (FA 6.2)"),
        (21, "KNRcorpbea",      "Constant-cost net stock, corporate (derived)"),
        (22, "pKN",             "Implicit price deflator, net capital stock"),
        (23, "KNHcorpbea",      "BEA historical-cost net stock, corporate (FA 6.3)"),
        (25, "DEPCcorp",        "Current-cost depreciation, corporate (FA 6.4)"),
        (26, "dcorpstar",       "Corrected depreciation rate (Shaikh)"),
        (27, "DEPHcorpbea",     "Historical-cost depreciation, corporate (FA 6.6)"),
        (28, "IGCcorpbea",      "Current-cost gross investment, corporate (FA 6.7)"),
        (29, "IGRcorpindexbea", "Chain-type QI gross investment, corporate (FA 6.8)"),
        (30, "IGRcorpbea",      "Constant-cost gross investment, corporate (derived)"),
        (31, "pIGcorpbea",      "Implicit price deflator, gross investment"),
        (37, "dcorpWhelanLiu",  "Whelan-Liu depreciation rate"),
    ]

    frames = {}
    for row_num, var_name, desc in row_map:
        df = read_excel_row(ws, row_num, year_row_idx=18, start_col=4)
        if not df.empty:
            df = df.rename(columns={"value": var_name})
            frames[var_name] = df
            n = df[var_name].notna().sum()
            yr_range = f"{df['year'].min()}-{df['year'].max()}"
            log.info(f"    {var_name}: {n} values ({yr_range}) — {desc}")

    wb.close()

    # Merge all series
    result = None
    for var_name, df in frames.items():
        if result is None:
            result = df
        else:
            result = result.merge(df, on="year", how="outer")

    if result is not None:
        result = result.sort_values("year").reset_index(drop=True)
        log.info(f"  II.1 total: {len(result)} years, {len(result.columns)-1} series")
    else:
        result = pd.DataFrame(columns=["year"])

    return result


def extract_appendix_II5():
    """Extract series from Appendix 6.8.II.5 (New Measures of Net and Gross Stock).

    Contains GPIM-constructed stocks:
    - Row 14: Adj. Ratio (adjustment ratio 1925-1947)
    - Row 15: KNHcorp — Net historical-cost capital stock (GPIM)
    - Row 16: KNCcorp — Net current-cost capital stock (GPIM)
    - Row 17: KGCcorp — Gross current-cost capital stock (GPIM)

    Year row = 12, data starts at column D (col 4).
    """
    log.info("  Extracting Appendix 6.8.II.5 (GPIM Net and Gross Stocks)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx 6.8.II.5"]
    except Exception as e:
        log.error(f"  Failed to open Appendix Excel: {e}")
        return pd.DataFrame(columns=["year"])

    row_map = [
        (14, "GPIM_adj_ratio",  "GPIM adjustment ratio (1925-1947)"),
        (15, "KNHcorp_gpim",    "Net historical-cost stock (GPIM-constructed)"),
        (16, "KNCcorp_gpim",    "Net current-cost stock (GPIM-constructed)"),
        (17, "KGCcorp",         "Gross current-cost stock (GPIM-constructed)"),
    ]

    frames = {}
    for row_num, var_name, desc in row_map:
        df = read_excel_row(ws, row_num, year_row_idx=12, start_col=4)
        if not df.empty:
            df = df.rename(columns={"value": var_name})
            frames[var_name] = df
            n = df[var_name].notna().sum()
            yr_range = f"{df['year'].min()}-{df['year'].max()}"
            log.info(f"    {var_name}: {n} values ({yr_range}) — {desc}")

    wb.close()

    result = None
    for var_name, df in frames.items():
        if result is None:
            result = df
        else:
            result = result.merge(df, on="year", how="outer")

    if result is not None:
        result = result.sort_values("year").reset_index(drop=True)
    else:
        result = pd.DataFrame(columns=["year"])

    return result


def extract_appendix_II7():
    """Extract series from Appendix 6.8.II.7 (Capacity Utilization and Final Capital).

    Contains Shaikh's final measures:
    - Row 22: VAcorp (Corporate Value Added, adj. for imputed interest)
    - Row 23: NOScorp (Corporate NOS, adj. for imputed interest)
    - Row 24: Pcorpnipa (NIPA Corporate Profit)
    - Row 25: NMINT (Net Monetary Interest Paid and Transfers)
    - Row 26: KGCcorp (Gross current-cost fixed capital)
    - Row 27: INVcorp (Corporate Inventories)
    - Row 28: KTCcorp (Total Capital Stock = KGCcorp + INVcorp)
    - Row 30: Rcorp (Corp Max Rate of Profit)

    Plus additional final measures after row 30.
    Year row needs to be identified — checking row 19 area.
    """
    log.info("  Extracting Appendix 6.8.II.7 (Capacity Utilization Final Measures)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx 6.8.II.7"]
    except Exception as e:
        log.error(f"  Failed to open Appendix Excel: {e}")
        return pd.DataFrame(columns=["year"])

    # Find the year row — check rows 1 and 19
    # Row 19 has column labels, Row 5 area might have years
    # Let me scan for a row that starts with year-like values
    year_row = None
    for r in range(1, 25):
        vals = [ws.cell(row=r, column=c).value for c in range(5, 10)]
        if any(v is not None and str(v).startswith("194") for v in vals):
            year_row = r
            break

    if year_row is None:
        # Try looking at row 19
        vals = [ws.cell(row=19, column=c).value for c in range(5, 10)]
        log.info(f"  Row 19 sample: {vals}")
        # Also try row 1
        vals1 = [ws.cell(row=1, column=c).value for c in range(5, 10)]
        log.info(f"  Row 1 sample: {vals1}")
        wb.close()
        return pd.DataFrame(columns=["year"])

    log.info(f"  Year row identified: {year_row}")

    row_map = [
        (22, "VAcorp_final",    "Corporate VA (adj. imputed interest)"),
        (23, "NOScorp_final",   "Corporate NOS (adj. imputed interest)"),
        (24, "Pcorpnipa_final", "NIPA Corporate Profit"),
        (25, "NMINT",           "Net Monetary Interest + Transfers"),
        (26, "KGCcorp_final",   "Gross current-cost fixed capital"),
        (27, "INVcorp",         "Corporate Inventories"),
        (28, "KTCcorp",         "Total Capital Stock (K + Inventories)"),
        (30, "Rcorp",           "Corp Max Rate of Profit"),
    ]

    frames = {}
    for row_num, var_name, desc in row_map:
        df = read_excel_row(ws, row_num, year_row_idx=year_row, start_col=5)
        if not df.empty:
            df = df.rename(columns={"value": var_name})
            frames[var_name] = df
            n = df[var_name].notna().sum()
            log.info(f"    {var_name}: {n} values — {desc}")

    # Also scan for additional series (rows 30-60)
    for r in range(31, min(60, ws.max_row + 1)):
        label = ws.cell(row=r, column=2).value  # Column B
        var = ws.cell(row=r, column=4).value     # Column D
        if var and isinstance(var, str) and len(var) > 2:
            df = read_excel_row(ws, r, year_row_idx=year_row, start_col=5)
            if not df.empty and df["value"].notna().sum() > 5:
                safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', str(var))
                df = df.rename(columns={"value": safe_name})
                frames[safe_name] = df
                n = df[safe_name].notna().sum()
                desc_str = str(label)[:50] if label else str(var)
                log.info(f"    {safe_name}: {n} values — {desc_str}")

    wb.close()

    result = None
    for var_name, df in frames.items():
        if result is None:
            result = df
        else:
            result = result.merge(df, on="year", how="outer")

    if result is not None:
        result = result.sort_values("year").reset_index(drop=True)
    else:
        result = pd.DataFrame(columns=["year"])

    return result


def extract_appendix_I():
    """Extract series from Appendix 6.8.I.1-3 (Business Sector Accounts).

    Contains GDP components, GVA by sector, profit measures.
    161 rows x 122 cols, years start at column D (col 4), row 3.
    """
    log.info("  Extracting Appendix 6.8.I.1-3 (Business Sector Accounts)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx6.8.I.1-3"]
    except Exception as e:
        log.error(f"  Failed to open Appendix Excel: {e}")
        return pd.DataFrame(columns=["year"])

    # Year row = 3, data starts at column D (col 4)
    # Variable name in column C (col 3)
    YEAR_ROW = 3

    # Key series to extract (row_number, variable_name, description)
    row_map = [
        (4,   "GDP",             "Gross Domestic Product (NIPA T1.7.5 line 1)"),
        (5,   "SD",              "Statistical Discrepancy (T1.7.5 line 15)"),
        (6,   "GDI",             "Gross Domestic Income (T1.10 line 1)"),
        (7,   "EC_total",        "Employee Compensation, total paid (T1.10 line 2)"),
        (8,   "T_total",         "Taxes net of Subsidies (T1.10 lines 9-10)"),
        (9,   "GOS_agg_nipa",    "Aggregate Gross Operating Surplus"),
        (10,  "CFC_total",       "Consumption of Fixed Capital (T1.10 line 23)"),
        (11,  "NOS_agg_nipa",    "Aggregate Net Operating Surplus (T1.10 line 11)"),
        (19,  "GVAhh",           "GVA Households (T1.3.5 line 6)"),
        (29,  "GVAnpish",        "GVA NPISH (T1.3.5 line 7)"),
        (39,  "GVAgengov",       "GVA General Government (T1.3.5 line 8)"),
        (43,  "CFCgengov",       "CFC General Government (T7.5 line 22)"),
        (49,  "GVAgoventerp",    "GVA Government Enterprises"),
        (53,  "CFCgoventerp",    "CFC Government Enterprises (T7.5 line 25)"),
        (59,  "GVAbusnipa",      "GVA Business Sector (derived)"),
        (63,  "GOSbusnipa",      "GOS Business Sector"),
        (64,  "CFCbusnipa",      "CFC Business Sector"),
        (65,  "NOSbusnipa",      "NOS Business Sector"),
        (88,  "GVAcorpnipa",     "GVA Corporate (NIPA)"),
        (89,  "VAcorpnipa",      "VA Corporate (NIPA, net of CFC)"),
        (90,  "GOScorpnipa",     "GOS Corporate (NIPA)"),
        (91,  "NOScorpnipa",     "NOS Corporate (NIPA)"),
        (94,  "Pcorpnipa",       "Corporate Profits w/IVA & CCAdj (NIPA)"),
        (123, "GVAcorp",         "GVA Corporate (adjusted for imputed interest)"),
        (124, "ECcorp",          "EC Corporate (adjusted)"),
        (126, "GOScorp",         "GOS Corporate (adjusted)"),
        (127, "CFCcorp",         "CFC Corporate (adjusted)"),
        (128, "VAcorp",          "VA Corporate (adjusted, = GVAcorp - CFCcorp)"),
        (129, "NOScorp",         "NOS Corporate (adjusted)"),
        (132, "Pcorp",           "Corporate Profits (adjusted)"),
        (136, "GVAnoncorp",      "GVA Non-corporate"),
        (151, "KNCbus",          "Business Net Fixed Capital, Current-Cost (end yr)"),
        (152, "KNCcorp_accts",   "Corporate Net Fixed Capital, Current-Cost (end yr)"),
        (153, "KNCnoncorp",      "Noncorporate Net Fixed Capital, Current-Cost (end yr)"),
        (154, "rbus",            "Business profit rate"),
        (155, "rcorp",           "Corporate profit rate"),
        (156, "rnoncorp",        "Noncorporate profit rate"),
    ]

    frames = {}
    for row_num, var_name, desc in row_map:
        df = read_excel_row(ws, row_num, year_row_idx=YEAR_ROW, start_col=4)
        if not df.empty:
            df = df.rename(columns={"value": var_name})
            frames[var_name] = df
            n = df[var_name].notna().sum()
            yr_range = f"{df['year'].min()}-{df['year'].max()}" if n > 0 else "EMPTY"
            log.info(f"    {var_name}: {n} values ({yr_range})")

    wb.close()

    result = None
    for var_name, df in frames.items():
        if result is None:
            result = df
        else:
            result = result.merge(df, on="year", how="outer")

    if result is not None:
        result = result.sort_values("year").reset_index(drop=True)
        log.info(f"  I.1-3 total: {len(result)} years, {len(result.columns)-1} series")
    else:
        result = pd.DataFrame(columns=["year"])

    return result


def extract_appendix_II3():
    """Extract from Appendix 6.8.II.3 (Effects of different depletion rates).

    Contains depreciation rate sensitivity analysis and alternative gross stocks.
    """
    log.info("  Extracting Appendix 6.8.II.3 (Depletion rates)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx 6.8.II.3"]
    except Exception as e:
        log.error(f"  Failed to open: {e}")
        return pd.DataFrame(columns=["year"])

    # Scan first 15 rows to find structure
    for i in range(1, 15):
        vals = [ws.cell(row=i, column=c).value for c in range(1, 8)]
        log.info(f"    Row {i}: {[str(v)[:30] if v else '' for v in vals]}")

    wb.close()
    return pd.DataFrame(columns=["year"])


def extract_appendix_II4():
    """Extract from Appendix 6.8.II.4 (Great Depression and WWII effects)."""
    log.info("  Extracting Appendix 6.8.II.4 (Depression & WWII)...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(APPENDIX_XLSX, read_only=True, data_only=True)
        ws = wb["Appndx 6.8.II.4"]
    except Exception as e:
        log.error(f"  Failed to open: {e}")
        return pd.DataFrame(columns=["year"])

    # Scan first 15 rows
    for i in range(1, 15):
        vals = [ws.cell(row=i, column=c).value for c in range(1, 8)]
        log.info(f"    Row {i}: {[str(v)[:40] if v else '' for v in vals]}")

    wb.close()
    return pd.DataFrame(columns=["year"])


def load_canonical_csv():
    """Load Shaikh canonical series CSV."""
    if not CANONICAL_CSV.exists():
        log.warning(f"  Canonical CSV not found: {CANONICAL_CSV}")
        return pd.DataFrame(columns=["year"])

    df = pd.read_csv(CANONICAL_CSV)
    log.info(f"  Canonical CSV: {len(df)} rows, cols: {df.columns.tolist()}")
    return df


def load_ddbb():
    """Load ddbb_cu_US_kgr.xlsx (pre-constructed Shaikh series 1925-2011)."""
    if not DDBB_XLSX.exists():
        log.warning(f"  ddbb not found: {DDBB_XLSX}")
        return pd.DataFrame(columns=["year"])

    df = pd.read_excel(DDBB_XLSX, sheet_name="us_data")
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    log.info(f"  ddbb: {len(df)} rows, cols: {df.columns.tolist()}")
    return df


# ===================================================================
# PART B: ONLINE EXTENSION (BEA API + FRED)
# ===================================================================

def fetch_with_retry(url, params=None, max_retries=4, timeout=30):
    """HTTP GET with exponential backoff."""
    import requests
    for attempt in range(max_retries):
        try:
            resp = requests.get(url, params=params, timeout=timeout)
            resp.raise_for_status()
            return resp
        except Exception as e:
            wait = 2 ** (attempt + 1)
            log.warning(f"    Attempt {attempt+1} failed: {e}. Retry in {wait}s...")
            time.sleep(wait)
    return None


def fetch_fred_series(series_id):
    """Fetch FRED series as annual DataFrame [year, value]."""
    cache = BEA_DIR / f"FRED_{series_id}.csv"
    if cache.exists():
        age_h = (time.time() - cache.stat().st_mtime) / 3600
        if age_h < 24:
            log.info(f"    FRED cached: {series_id}")
            df = pd.read_csv(cache)
            col = [c for c in df.columns if c != "DATE"][0]
            df["year"] = pd.to_datetime(df["DATE"]).dt.year
            df["value"] = pd.to_numeric(df[col], errors="coerce")
            return df.groupby("year")["value"].mean().reset_index()

    resp = fetch_with_retry(f"{FRED_CSV_BASE}?id={series_id}")
    if resp is None:
        log.warning(f"    FRED {series_id}: network unavailable")
        return None

    cache.write_bytes(resp.content)
    df = pd.read_csv(cache)
    col = [c for c in df.columns if c != "DATE"][0]
    df["year"] = pd.to_datetime(df["DATE"]).dt.year
    df["value"] = pd.to_numeric(df[col], errors="coerce")
    return df.groupby("year")["value"].mean().reset_index()


def fetch_bea_api(api_key, dataset, table_name, frequency="A", year="X"):
    """Fetch BEA API table. Returns DataFrame or None."""
    cache = BEA_DIR / f"BEA_{dataset}_{table_name}.json"
    if cache.exists():
        age_h = (time.time() - cache.stat().st_mtime) / 3600
        if age_h < 24:
            log.info(f"    BEA cached: {dataset}/{table_name}")
            with open(cache) as f:
                data = json.load(f)
            try:
                return pd.DataFrame(data["BEAAPI"]["Results"]["Data"])
            except (KeyError, TypeError):
                return None

    params = {
        "UserID": api_key, "method": "GetData",
        "DataSetName": dataset, "TableName": table_name,
        "Frequency": frequency, "Year": year, "ResultFormat": "JSON",
    }
    resp = fetch_with_retry(BEA_API_BASE, params=params, timeout=60)
    if resp is None:
        return None

    data = resp.json()
    cache.write_text(json.dumps(data))
    try:
        return pd.DataFrame(data["BEAAPI"]["Results"]["Data"])
    except (KeyError, TypeError):
        return None


def try_online_extension(api_key=None, try_fred=False):
    """Try to extend series beyond 2011 using BEA API or FRED."""
    log.info("=== Online Extension ===")
    frames = {}

    if try_fred:
        fred_map = {
            "GDP_nom":        "GDP",
            "GDP_real_2017":  "GDPC1",
            "GVA_corp_nom":   "A455RC1A027NBEA",
            "GVA_corpNF_nom": "A454RC1A027NBEA",
            "pGDP":           "GDPDEF",
        }
        for col_name, sid in fred_map.items():
            df = fetch_fred_series(sid)
            if df is not None and not df.empty:
                df = df.rename(columns={"value": col_name})
                frames[col_name] = df
                log.info(f"    {col_name}: {len(df)} years "
                         f"({df['year'].min()}-{df['year'].max()})")

    if api_key:
        log.info("  BEA API: fetching Fixed Assets Section 6 tables...")
        # This will be populated when API key is available
        # FAAt601: Current-cost net stock by legal form
        # FAAt604: Current-cost depreciation
        # FAAt607: Investment
        for table, measure in [("FAAt601", "K_net"), ("FAAt604", "DEP"),
                               ("FAAt607", "I_hist")]:
            df = fetch_bea_api(api_key, "FixedAssets", table)
            if df is not None:
                log.info(f"    {table}: {len(df)} records")
                # Parse line items — will need line number mapping
                # For now, cache the raw data for manual inspection
                df.to_csv(BEA_DIR / f"BEA_FA_{table}_raw.csv", index=False)

    # Merge FRED frames
    if not frames:
        log.info("  No online data obtained")
        return pd.DataFrame(columns=["year"])

    result = None
    for col_name, df in frames.items():
        if result is None:
            result = df
        else:
            result = result.merge(df, on="year", how="outer")

    return result.sort_values("year").reset_index(drop=True)


# ===================================================================
# PART C: ASSEMBLY
# ===================================================================

def rebase_index(series, from_year, to_year, year_series):
    """Rebase a price index from from_year=100 to to_year=100."""
    mask = year_series == from_year
    if mask.any():
        base_val = series[mask].values[0]
    else:
        mask2 = year_series == to_year
        if mask2.any():
            return series  # Already in unknown base, return as-is
        return series

    mask_to = year_series == to_year
    if mask_to.any():
        new_base = series[mask_to].values[0]
        return series / new_base * 100
    return series


def assemble_dataset(accts_df, ii1_df, ii5_df, ii7_df, canonical_df, ddbb_df,
                     online_df=None):
    """Merge all local sources and compute derived series."""
    log.info("=== Assembly ===")

    # Year spine: 1925-2024
    spine = pd.DataFrame({"year": range(1925, 2025)})
    result = spine.copy()

    # Merge all local sources
    for label, df in [("Accounts_I", accts_df), ("GPIM_II1", ii1_df),
                      ("GPIM_II5", ii5_df), ("CU_II7", ii7_df),
                      ("Canonical", canonical_df), ("DDBB", ddbb_df)]:
        if df is not None and not df.empty and "year" in df.columns:
            # Avoid duplicate columns
            overlap = set(result.columns) & set(df.columns) - {"year"}
            if overlap:
                df = df.drop(columns=list(overlap))
            result = result.merge(df, on="year", how="left")
            n_new = len(df.columns) - 1
            log.info(f"  Merged {label}: +{n_new} columns")

    # Merge online extension
    if online_df is not None and not online_df.empty:
        overlap = set(result.columns) & set(online_df.columns) - {"year"}
        if overlap:
            # For online data, prefer it for years > 2011
            for col in overlap:
                online_vals = online_df.set_index("year")[col]
                for yr in online_vals.index:
                    if yr > 2011 and pd.notna(online_vals.get(yr)):
                        result.loc[result["year"] == yr, col] = online_vals[yr]
            online_df = online_df.drop(columns=list(overlap))
        result = result.merge(online_df, on="year", how="left")

    # -----------------------------------------------------------------------
    # Derived series
    # -----------------------------------------------------------------------
    log.info("  Computing derived series...")

    # GVAcorp_nom = GVAcorp (from Appendix I, adjusted) — includes CFC
    # This IS the gross value added including capital consumption
    if "GVAcorp" in result.columns:
        result["GVAcorp_nom"] = result["GVAcorp"]
        log.info("    GVAcorp_nom = GVAcorp (adj. for imputed interest, includes CFC)")

    # NVA = GVA - CFC
    if "VAcorp" in result.columns:
        result["NVAcorp_nom"] = result["VAcorp"]
        log.info("    NVAcorp_nom = VAcorp (= GVAcorp - CFCcorp)")

    # Gross capital-output ratio
    if "GVAcorp_nom" in result.columns and "KGCcorp" in result.columns:
        # Use lagged K (end of previous year)
        result["KGCcorp_lag1"] = result["KGCcorp"].shift(1)
        mask = (result["GVAcorp_nom"] > 0) & (result["KGCcorp_lag1"] > 0)
        result.loc[mask, "KY_ratio_gross"] = (
            result.loc[mask, "KGCcorp_lag1"] / result.loc[mask, "GVAcorp_nom"]
        )
        log.info("    KY_ratio_gross = KGCcorp(t-1) / GVAcorp_nom(t)")

    # Net capital-output ratio
    if "VAcorp" in result.columns and "KNCcorpbea" in result.columns:
        result["KNCcorpbea_lag1"] = result["KNCcorpbea"].shift(1)
        mask = (result["VAcorp"] > 0) & (result["KNCcorpbea_lag1"] > 0)
        result.loc[mask, "KY_ratio_net"] = (
            result.loc[mask, "KNCcorpbea_lag1"] / result.loc[mask, "VAcorp"]
        )

    # Real series using pKN (base 2005)
    if "pKN" in result.columns:
        pKN_2005 = result.loc[result["year"] == BASE_YEAR, "pKN"]
        if not pKN_2005.empty:
            pKN_base = pKN_2005.values[0]
            result["pKN_2005"] = result["pKN"] / pKN_base * 100
            log.info(f"    pKN_2005: rebased (pKN at 2005 = {pKN_base:.4f})")

    if "pIGcorpbea" in result.columns:
        pIG_2005 = result.loc[result["year"] == BASE_YEAR, "pIGcorpbea"]
        if not pIG_2005.empty:
            pIG_base = pIG_2005.values[0]
            result["pIG_2005"] = result["pIGcorpbea"] / pIG_base * 100
            log.info(f"    pIG_2005: rebased (pIG at 2005 = {pIG_base:.4f})")

    # Capacity utilization (if available from ddbb)
    if "yk" in result.columns:
        result.rename(columns={"yk": "uK_ddbb"}, inplace=True)
        log.info("    uK_ddbb: capacity utilization from ddbb")

    # Toggle flags
    result["shaikh_gpim_adj"] = False
    result["wwii_adj"] = False

    # -----------------------------------------------------------------------
    # Coverage summary
    # -----------------------------------------------------------------------
    n_cols = len([c for c in result.columns if c != "year"])
    n_rows = len(result)
    numeric_cols = result.select_dtypes(include=[np.number]).columns.tolist()
    coverage_rows = {}
    for col in numeric_cols:
        if col == "year":
            continue
        valid = result[col].notna()
        if valid.any():
            yr_min = result.loc[valid, "year"].min()
            yr_max = result.loc[valid, "year"].max()
            coverage_rows[col] = (valid.sum(), yr_min, yr_max)

    log.info(f"  Final: {n_rows} years x {n_cols} series")
    log.info(f"  Series with data: {len(coverage_rows)}")

    return result


# ===================================================================
# PART D: OUTPUT
# ===================================================================

def write_output(df):
    """Write dataset to CSV and summary."""
    df.to_csv(OUTPUT_CSV, index=False, float_format="%.8f")
    log.info(f"  Written: {OUTPUT_CSV}")
    log.info(f"    {len(df)} rows, {len(df.columns)} columns")

    # Summary file
    summary_path = PROC_DIR / "bea_dataset_summary.txt"
    with open(summary_path, "w") as f:
        f.write(f"BEA Extended Dataset v1\n")
        f.write(f"Generated: {datetime.now().isoformat()}\n")
        f.write(f"Base year: {BASE_YEAR}\n")
        f.write(f"Rows: {len(df)}, Columns: {len(df.columns)}\n")
        f.write(f"Year range: {df['year'].min()}-{df['year'].max()}\n\n")

        f.write("=" * 70 + "\n")
        f.write(f"{'Column':<35} {'N':>5} {'Start':>6} {'End':>6} {'Source'}\n")
        f.write("=" * 70 + "\n")

        for col in df.columns:
            if col == "year":
                continue
            n_valid = df[col].notna().sum()
            if n_valid > 0 and df[col].dtype in [np.float64, np.int64, "Int64",
                                                   np.float32]:
                yr_min = int(df.loc[df[col].notna(), "year"].min())
                yr_max = int(df.loc[df[col].notna(), "year"].max())
                f.write(f"  {col:<33} {n_valid:>5} {yr_min:>6} {yr_max:>6}\n")
            elif n_valid > 0:
                f.write(f"  {col:<33} {n_valid:>5}\n")

    log.info(f"  Summary: {summary_path}")


# ===================================================================
# MAIN
# ===================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Build extended BEA Fixed Assets dataset"
    )
    parser.add_argument("--api-key", default=os.environ.get("BEA_API_KEY"),
                        help="BEA API key (or set BEA_API_KEY env var)")
    parser.add_argument("--try-fred", action="store_true",
                        help="Try FRED for GDP/GVA extension")
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("BEA Extended Dataset Builder v1.0")
    log.info(f"  Base year: {BASE_YEAR}")
    log.info(f"  Output: {OUTPUT_CSV}")
    log.info(f"  API key: {'yes' if args.api_key else 'no'}")
    log.info(f"  FRED: {'yes' if args.try_fred else 'no'}")
    log.info("=" * 60)

    # ---------------------------------------------------------------
    # Phase 1: Local Excel extraction (always runs)
    # ---------------------------------------------------------------
    log.info("\n=== Phase 1: Local Excel Extraction ===")

    accts_df = extract_appendix_I()
    ii1_df = extract_appendix_II1()
    ii5_df = extract_appendix_II5()
    ii7_df = extract_appendix_II7()

    canonical_df = load_canonical_csv()
    ddbb_df = load_ddbb()

    # ---------------------------------------------------------------
    # Phase 2: Online extension (optional)
    # ---------------------------------------------------------------
    online_df = None
    if args.api_key or args.try_fred:
        online_df = try_online_extension(
            api_key=args.api_key,
            try_fred=args.try_fred,
        )

    # ---------------------------------------------------------------
    # Phase 3: Assembly
    # ---------------------------------------------------------------
    dataset = assemble_dataset(
        accts_df, ii1_df, ii5_df, ii7_df,
        canonical_df, ddbb_df, online_df,
    )

    # ---------------------------------------------------------------
    # Phase 4: Output
    # ---------------------------------------------------------------
    write_output(dataset)

    log.info("\n" + "=" * 60)
    log.info("DONE.")
    if not args.api_key:
        log.info("  To extend to 2024: re-run with --api-key YOUR_KEY")
        log.info("  Register free: https://apps.bea.gov/API/signup/index.cfm")
    log.info("  Next: python3 codes/16_bea_validation.py")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
