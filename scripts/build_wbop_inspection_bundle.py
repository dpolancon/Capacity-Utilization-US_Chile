from __future__ import annotations

import logging
import math
import re
import textwrap
from dataclasses import dataclass
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "data" / "raw" / "other" / "NievasPiketty2025WBOPFinalSeries.xlsx"

OUTPUT_ROOT = REPO_ROOT / "output" / "wbop"
OUTPUT_PNG = OUTPUT_ROOT / "figures" / "png"
OUTPUT_PDF = OUTPUT_ROOT / "figures" / "pdf"
OUTPUT_LOGS = OUTPUT_ROOT / "logs"

PROCESSED_ROOT = REPO_ROOT / "data" / "processed"

VISUAL_MD = OUTPUT_ROOT / "wbop_visual_inspection.md"
VISUAL_INVENTORY_CSV = OUTPUT_ROOT / "wbop_visual_inventory.csv"
EXCLUDED_SERIES_CSV = OUTPUT_ROOT / "wbop_excluded_series.csv"
MISSINGNESS_CSV = OUTPUT_ROOT / "wbop_missingness_summary.csv"
DUPLICATE_CSV = OUTPUT_ROOT / "wbop_duplicate_check.csv"
ENTITY_MAPPING_CSV = OUTPUT_ROOT / "wbop_entity_mapping_detected.csv"
BUILD_LOG = OUTPUT_LOGS / "build_log.txt"

CHILE_US_PANEL_CSV = PROCESSED_ROOT / "wbop_chile_us_panel.csv"
CHILE_US_GUIDE_MD = PROCESSED_ROOT / "wbop_chile_us_user_guide.md"
VARIABLE_INVENTORY_XLSX = PROCESSED_ROOT / "wbop_variable_inventory.xlsx"

HISTORICAL_MARKERS = [
    1914,
    1917,
    1929,
    1932,
    1939,
    1944,
    1947,
    1957,
    1964,
    1970,
    1972,
    1975,
    1978,
    1982,
    1987,
    1990,
    1997,
    2001,
    2008,
    2011,
    2014,
    2019,
]

TARGET_ENTITIES = {
    "United States": ["United States", "USA", "US"],
    "United Kingdom": ["United Kingdom", "Britain", "GB", "UK", "Great Britain"],
    "Chile": ["Chile", "CL"],
    "Latin America": ["Latin America", "LATAM"],
}

AUXILIARY_SHEETS = {"ReadMe", "WBOP"}


@dataclass
class SheetProfile:
    sheet_name: str
    dimensions: str
    header_structure: str
    likely_identifier_columns: str
    likely_time_column: str
    numeric_columns: int
    metadata_columns: str
    sheet_role: str
    title: str
    header_row_index: int | None
    year_start_row_index: int | None
    time_column_name: str | None
    min_year: int | None
    max_year: int | None


def configure_logging(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("wbop_bundle")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(stream_handler)

    return logger


def ensure_directories() -> None:
    for path in [OUTPUT_ROOT, OUTPUT_PNG, OUTPUT_PDF, OUTPUT_LOGS, PROCESSED_ROOT]:
        path.mkdir(parents=True, exist_ok=True)


def clean_string(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def canonicalize_label(value: str) -> str:
    value = clean_string(value)
    value = re.sub(r"[/(),.%+-]", " ", value)
    value = re.sub(r"\s+", " ", value).strip().lower()
    return value


def sanitize_filename(text: str) -> str:
    text = canonicalize_label(text)
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text[:140] or "series"


def parse_year(value: object) -> int | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        year = int(round(float(value)))
        if 1500 <= year <= 2100:
            return year
        return None

    text = clean_string(value)
    if not text:
        return None

    match = re.search(r"(?<!\d)(1[5-9]\d{2}|20\d{2}|2100)(?!\d)", text)
    if match:
        return int(match.group(0))
    return None


def find_header_row(raw_df: pd.DataFrame) -> int | None:
    best_row = None
    best_score = -1
    max_scan = min(10, len(raw_df))
    for row_idx in range(max_scan):
        row = raw_df.iloc[row_idx]
        nonempty_after_first = sum(clean_string(x) != "" for x in row.iloc[1:])
        string_like_after_first = sum(
            clean_string(x) != "" and parse_year(x) is None for x in row.iloc[1:]
        )
        score = nonempty_after_first * 10 + string_like_after_first
        if score > best_score and nonempty_after_first >= 4:
            best_score = score
            best_row = row_idx
    return best_row


def detect_year_start_row(raw_df: pd.DataFrame, header_row: int | None) -> int | None:
    if header_row is None:
        return None
    for row_idx in range(header_row + 1, len(raw_df)):
        if parse_year(raw_df.iat[row_idx, 0]) is not None:
            return row_idx
    return None


def nonempty_dimensions(raw_df: pd.DataFrame) -> tuple[int, int]:
    nonempty_rows = int(raw_df.dropna(how="all").shape[0])
    nonempty_cols = int(raw_df.dropna(axis=1, how="all").shape[1])
    return nonempty_rows, nonempty_cols


def inspect_sheet(sheet_name: str, raw_df: pd.DataFrame) -> SheetProfile:
    nonempty_rows, nonempty_cols = nonempty_dimensions(raw_df)
    title = clean_string(raw_df.iat[0, 0]) if not raw_df.empty else ""

    if sheet_name in AUXILIARY_SHEETS or nonempty_cols <= 2:
        role = "auxiliary"
        header_structure = "Free-text auxiliary sheet, not a numeric panel source."
        id_cols = "None detected as tabular entity identifiers."
        time_col = "None detected."
        numeric_cols = 0
        metadata_cols = "All content appears descriptive."
        return SheetProfile(
            sheet_name=sheet_name,
            dimensions=f"{raw_df.shape[0]}x{raw_df.shape[1]}",
            header_structure=header_structure,
            likely_identifier_columns=id_cols,
            likely_time_column=time_col,
            numeric_columns=numeric_cols,
            metadata_columns=metadata_cols,
            sheet_role=role,
            title=title,
            header_row_index=None,
            year_start_row_index=None,
            time_column_name=None,
            min_year=None,
            max_year=None,
        )

    header_row = find_header_row(raw_df)
    year_start_row = detect_year_start_row(raw_df, header_row)
    header_values = raw_df.iloc[header_row].tolist() if header_row is not None else []
    stripped_headers = [clean_string(x) for x in header_values]

    time_column_name = stripped_headers[0] or "derived_year_from_first_column"
    entity_count = sum(1 for x in stripped_headers[1:] if x)
    min_year = None
    max_year = None
    if year_start_row is not None:
        year_values = raw_df.iloc[year_start_row:, 0].map(parse_year).dropna().astype(int)
        if not year_values.empty:
            min_year = int(year_values.min())
            max_year = int(year_values.max())

    header_structure = (
        f"Title row in row 1; likely entity header row in row {header_row + 1}; "
        f"years begin in row {year_start_row + 1 if year_start_row is not None else 'NA'}."
    )
    id_cols = (
        f"Columns 2-{entity_count + 1} contain entity labels/codes "
        f"(sample: {', '.join(stripped_headers[1:6])})."
    )
    time_col = (
        f"First column parsed as {time_column_name}; values are annual years "
        f"covering {min_year}-{max_year}."
    )
    metadata_cols = "No metadata columns detected inside the tabular body beyond title/header rows."
    role = "main_panel_source"

    return SheetProfile(
        sheet_name=sheet_name,
        dimensions=f"{raw_df.shape[0]}x{raw_df.shape[1]}",
        header_structure=header_structure,
        likely_identifier_columns=id_cols,
        likely_time_column=time_col,
        numeric_columns=entity_count,
        metadata_columns=metadata_cols,
        sheet_role=role,
        title=title,
        header_row_index=header_row,
        year_start_row_index=year_start_row,
        time_column_name=time_column_name,
        min_year=min_year,
        max_year=max_year,
    )


def normalize_data_sheet(raw_df: pd.DataFrame, profile: SheetProfile, source_sheet: str) -> pd.DataFrame:
    if profile.header_row_index is None or profile.year_start_row_index is None:
        raise ValueError(f"Sheet {source_sheet} does not have a detected data structure.")

    header_values = [clean_string(x) for x in raw_df.iloc[profile.header_row_index].tolist()]
    header_values[0] = "year"

    body = raw_df.iloc[profile.year_start_row_index :, : len(header_values)].copy()
    body.columns = header_values
    body["year"] = body["year"].map(parse_year)
    body = body[body["year"].notna()].copy()
    body["year"] = body["year"].astype(int)

    entity_columns = [col for col in body.columns if col != "year" and clean_string(col)]
    for col in entity_columns:
        body[col] = pd.to_numeric(body[col], errors="coerce")

    body = body.sort_values("year").reset_index(drop=True)
    body.attrs["title"] = profile.title
    body.attrs["source_sheet"] = source_sheet
    return body


def detect_target_entities(
    all_entity_labels: Iterable[str],
    wbop_text: str,
    logger: logging.Logger,
) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    observed = sorted({clean_string(x) for x in all_entity_labels if clean_string(x)})
    observed_lookup = {canonicalize_label(label): label for label in observed}

    mapping_rows: list[dict[str, str]] = []
    detected_map: dict[str, str] = {}
    manual_review_items: list[str] = []

    wbop_text_norm = canonicalize_label(wbop_text)

    for standardized_label, candidates in TARGET_ENTITIES.items():
        matched = None
        detection_source = "data_sheet_headers"
        for candidate in candidates:
            normalized_candidate = canonicalize_label(candidate)
            if normalized_candidate in observed_lookup:
                matched = observed_lookup[normalized_candidate]
                break

        if matched is None:
            close = get_close_matches(
                canonicalize_label(standardized_label),
                list(observed_lookup.keys()),
                n=3,
                cutoff=0.5,
            )
            if close:
                manual_review_items.append(
                    f"{standardized_label}: no exact header match; closest header labels were "
                    + ", ".join(observed_lookup[item] for item in close)
                )
            if canonicalize_label(standardized_label) in wbop_text_norm:
                detection_source = "wbop_sheet_only"
            matched = ""

        if standardized_label == "United Kingdom" and "britain" in wbop_text_norm:
            manual_review_items.append(
                "WBOP coverage sheet uses 'Britain' while numeric data sheets use 'GB'."
            )
        if standardized_label == "United States" and "usa" in wbop_text_norm:
            manual_review_items.append(
                "WBOP coverage sheet uses 'USA' while numeric data sheets use 'US'."
            )
        if standardized_label == "Chile" and "chile" in wbop_text_norm and matched == "CL":
            manual_review_items.append(
                "WBOP coverage sheet spells out 'Chile' while numeric data sheets use 'CL'."
            )

        if matched:
            detected_map[standardized_label] = matched

        entity_type_guess = "region" if standardized_label == "Latin America" else "country"
        mapping_rows.append(
            {
                "original_label": matched,
                "standardized_label": standardized_label,
                "entity_type_guess": entity_type_guess,
                "included_in_visual_bundle": "yes" if matched else "no",
                "included_in_chile_us_panel": (
                    "yes" if standardized_label in {"United States", "Chile"} and matched else "no"
                ),
                "detection_source": detection_source,
                "detection_notes": (
                    f"Detected from numeric data-sheet headers as '{matched}'."
                    if matched
                    else "No exact data-sheet header match detected."
                ),
            }
        )

    logger.info("Detected target entity mapping: %s", detected_map)

    deduped_manual_review = []
    seen = set()
    for item in manual_review_items:
        if item not in seen:
            deduped_manual_review.append(item)
            seen.add(item)

    mapping_df = pd.DataFrame(mapping_rows)
    return mapping_df, detected_map, deduped_manual_review


def build_missingness_summary(
    filtered_long: pd.DataFrame,
    plotted_flags: dict[str, bool],
) -> pd.DataFrame:
    rows = []
    grouped = filtered_long.groupby(["source_sheet", "original_variable_name", "standardized_entity"], dropna=False)
    for (source_sheet, variable_name, standardized_entity), grp in grouped:
        nonmissing = grp["value"].notna().sum()
        total = len(grp)
        missing = total - nonmissing
        nonmissing_years = grp.loc[grp["value"].notna(), "year"]
        constant_series = (
            grp.loc[grp["value"].notna(), "value"].nunique(dropna=True) <= 1 if nonmissing > 0 else False
        )
        rows.append(
            {
                "source_sheet": source_sheet,
                "original_variable_name": variable_name,
                "standardized_entity": standardized_entity,
                "raw_entity_label": grp["raw_entity_label"].iloc[0],
                "total_obs_post_1900": total,
                "nonmissing_obs_post_1900": int(nonmissing),
                "missing_obs_post_1900": int(missing),
                "missingness_share": round(missing / total, 6) if total else None,
                "min_year_available": int(nonmissing_years.min()) if not nonmissing_years.empty else None,
                "max_year_available": int(nonmissing_years.max()) if not nonmissing_years.empty else None,
                "constant_series_yes_no": "yes" if constant_series else "no",
                "plotted_variable_yes_no": "yes" if plotted_flags.get(source_sheet, False) else "no",
            }
        )
    return pd.DataFrame(rows).sort_values(
        ["source_sheet", "standardized_entity"]
    ).reset_index(drop=True)


def write_variable_inventory_workbook(
    inventory_df: pd.DataFrame,
    summary_rows: list[list[object]],
    output_path: Path,
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "variable_inventory"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(list(inventory_df.columns))
    for row in inventory_df.itertuples(index=False):
        ws.append(list(row))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 48)

    summary_ws = workbook.create_sheet("summary")
    for row in summary_rows:
        summary_ws.append(row)
    summary_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = summary_ws.dimensions
    for idx, column in enumerate(summary_ws.columns, start=1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        summary_ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)

    workbook.save(output_path)


def format_percent(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "NA"
    return f"{value:.1%}"


def format_year_range(min_year: int | None, max_year: int | None) -> str:
    if min_year is None or max_year is None:
        return "NA"
    return f"{min_year}-{max_year}"


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    if df.empty:
        return "_No rows._"

    columns = list(df.columns)
    display_df = df.copy()
    for col in columns:
        display_df[col] = display_df[col].map(lambda x: "" if pd.isna(x) else str(x))

    widths: list[int] = []
    for col in columns:
        max_value_width = int(display_df[col].map(len).max()) if not display_df.empty else 0
        widths.append(max(len(str(col)), max_value_width))

    header = "| " + " | ".join(str(col).ljust(width) for col, width in zip(columns, widths)) + " |"
    separator = "| " + " | ".join("-" * width for width in widths) + " |"
    body = []
    for _, row in display_df.iterrows():
        body.append(
            "| " + " | ".join(str(row[col]).ljust(widths[idx]) for idx, col in enumerate(columns)) + " |"
        )
    return "\n".join([header, separator] + body)


def write_visual_markdown(
    output_path: Path,
    workbook_profiles: list[SheetProfile],
    used_sheets: list[str],
    entity_mapping_df: pd.DataFrame,
    detected_entities: dict[str, str],
    visual_inventory_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    missingness_df: pd.DataFrame,
    manual_review_items: list[str],
    duplicate_df: pd.DataFrame,
    overall_year_range: tuple[int | None, int | None],
) -> None:
    lines: list[str] = []
    lines.append("# WBOP dataset inspection bundle")
    lines.append("")
    lines.append("## 1. Purpose")
    lines.append("")
    lines.append(
        "This bundle provides a descriptive inspection of the Nievas–Piketty WBOP dataset for Chapter 2."
    )
    lines.append(f"Source workbook: `{SOURCE_WORKBOOK}`.")
    lines.append(
        "The visual inspection focuses on United States, United Kingdom, Chile, and Latin America from 1900 onward."
    )
    lines.append("")
    lines.append("## 2. Workbook audit")
    lines.append("")
    lines.append(
        "Workbook sheets: "
        + ", ".join(f"`{profile.sheet_name}`" for profile in workbook_profiles)
        + "."
    )
    lines.append(
        "Used numeric data sheets: "
        + ", ".join(f"`{sheet}`" for sheet in used_sheets)
        + "."
    )
    lines.append(
        "Detected entity representation: wide columns across the data sheets, with country/regional labels in the header row."
    )
    lines.append(
        "Detected time variable: the first column is annual year. In most sheets the first-column header is blank and was conservatively treated as a derived year field; in `C1e` and `C1f` the first-column header is explicitly `year`."
    )
    lines.append("Dataset structure: wide format by sheet, with one numeric variable per sheet.")
    lines.append(
        f"Overall year coverage across used sheets: {format_year_range(*overall_year_range)}."
    )
    duplicate_issues = duplicate_df.loc[duplicate_df["duplicate_count"] > 0]
    lines.append(
        "Integrity handling: duplicates were checked both in sheet-level year rows and in the normalized entity-year-variable form. "
        + ("Non-zero duplicates were detected and logged." if not duplicate_issues.empty else "No non-zero duplicates were detected.")
    )
    lines.append("")
    lines.append("## 3. Entity coverage")
    lines.append("")

    entity_view = entity_mapping_df[
        [
            "standardized_label",
            "original_label",
            "entity_type_guess",
            "included_in_visual_bundle",
            "included_in_chile_us_panel",
        ]
    ].copy()
    lines.append(dataframe_to_markdown(entity_view))
    lines.append("")
    if manual_review_items:
        lines.append("Manual-review notes:")
        for item in manual_review_items:
            lines.append(f"- {item}")
    else:
        lines.append("No unresolved entity-label ambiguities remained after detection.")
    lines.append("")
    lines.append("## 4. Variable coverage")
    lines.append("")
    total_variables = len(used_sheets)
    total_plotted = int((visual_inventory_df["plotted_yes_no"] == "yes").sum())
    total_excluded = len(excluded_df)
    lines.append(f"Total variables inspected: {total_variables}.")
    lines.append(f"Total usable numeric variables: {total_plotted}.")
    lines.append(f"Total plotted: {total_plotted}.")
    lines.append(f"Total excluded: {total_excluded}.")
    lines.append("")
    compact_table = visual_inventory_df[
        [
            "source_sheet",
            "original_variable_name",
            "entities_available",
            "min_year",
            "max_year",
            "missingness_share",
        ]
    ].copy()
    compact_table["missingness_share"] = compact_table["missingness_share"].map(format_percent)
    lines.append(dataframe_to_markdown(compact_table))
    lines.append("")
    lines.append("## 5. Visual inspection")
    lines.append("")

    for row in visual_inventory_df.itertuples(index=False):
        lines.append(f"### {row.original_variable_name}")
        lines.append("")
        lines.append(
            f"Diagnostic note: entities available = {row.entities_available}; year range = {row.min_year}-{row.max_year}; "
            f"missingness share = {format_percent(row.missingness_share)}; source sheet = `{row.source_sheet}`."
        )
        lines.append("")
        lines.append(f"![{row.original_variable_name}]({Path(row.png_path).as_posix()})")
        lines.append("")
        lines.append(f"`{row.png_path}`")
        lines.append("")
        lines.append(f"`{row.pdf_path}`")
        lines.append("")

    lines.append("## 6. Descriptive synthesis")
    lines.append("")
    if visual_inventory_df.empty:
        lines.append("No usable post-1900 series were available for the requested entities.")
    else:
        avg_missingness = visual_inventory_df["missingness_share"].mean()
        fully_covered = int((visual_inventory_df["missingness_share"] == 0).sum())
        stock_vars = visual_inventory_df[
            visual_inventory_df["source_sheet"].isin(["G1a", "G1b", "G1c"])
        ]["original_variable_name"].tolist()
        lines.append(
            f"The inspection covers {total_plotted} plotted variables from {visual_inventory_df['min_year'].min()} to {visual_inventory_df['max_year'].max()} after the 1900 filter. "
            f"{fully_covered} variables have zero post-1900 missingness across the four requested entities, and the average missingness share across plotted variables is {format_percent(avg_missingness)}."
        )
        lines.append(
            "The bundle separates one level series in current millions of dollars from a larger block of percentage-of-GDP flow and stock series. "
            "That distinction matters visually: the GDP sheet tracks long-run scale divergence, while the trade, income, transfer, current-account, and foreign-wealth sheets emphasize oscillation, sign changes, and discrete breaks."
        )
        if stock_vars:
            lines.append(
                "The external stock sheets (`G1a`, `G1b`, `G1c`) show the sharpest swings and level shifts in the bundle. "
                "The flow sheets are generally tighter, but they still show visible breaks around interwar, postwar, 1970s, early-1980s, and post-2008 intervals."
            )
        lines.append(
            "Coverage remains uneven across variables rather than across the selected entities alone. "
            "The figures identify where continuity is strong enough for later substantive work and where semantic cleaning or closer source validation should come first."
        )
    lines.append("")
    lines.append("## 7. Audit appendix")
    lines.append("")
    if excluded_df.empty:
        lines.append("No numeric variables from the used data sheets were excluded from the visual bundle.")
    else:
        lines.append(dataframe_to_markdown(excluded_df))
    lines.append("")
    if manual_review_items:
        lines.append("Remaining ambiguities or relabeling items:")
        for item in manual_review_items:
            lines.append(f"- {item}")
    else:
        lines.append("No unresolved ambiguities remained after inspection.")
    lines.append("")
    lines.append("Variables that may need later semantic cleaning:")
    lines.append(
        "- Country and region headers are a mix of spelled-out aggregates and short codes. The visual bundle standardizes only the requested four entities, leaving the broader codebook to a later pass."
    )
    lines.append(
        "- Some sheets normalize region names slightly differently (`North America/ Oceania` vs `North America Oceania`; `Subsaharan Africa` vs `Sub-Saharan Africa`)."
    )
    lines.append("")
    lines.append("## 8. Processed Chile–United States outputs")
    lines.append("")
    lines.append(f"Processed panel: `{CHILE_US_PANEL_CSV}`.")
    lines.append(f"User guide: `{CHILE_US_GUIDE_MD}`.")
    lines.append(f"Variable inventory workbook: `{VARIABLE_INVENTORY_XLSX}`.")
    lines.append(
        "The processed panel keeps one row per standardized entity-year and preserves the original variable names as wide data columns. "
        "That choice matches the workbook’s one-sheet-per-variable structure and avoids an unnecessary long-to-wide-to-long cycle."
    )

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_chile_us_user_guide(
    output_path: Path,
    used_sheets: list[str],
    detected_entities: dict[str, str],
    overall_year_range: tuple[int | None, int | None],
    duplicate_df: pd.DataFrame,
    missingness_df: pd.DataFrame,
    variable_inventory_df: pd.DataFrame,
) -> None:
    duplicates_found = duplicate_df.loc[duplicate_df["duplicate_count"] > 0]
    ch_us_missing = missingness_df[
        missingness_df["standardized_entity"].isin(["Chile", "United States"])
    ].copy()
    ch_us_missing["missingness_share"] = ch_us_missing["missingness_share"].map(format_percent)

    lines = [
        "# WBOP Chile–United States processed panel",
        "",
        "## Scope",
        "",
        "This processed file contains all available WBOP series for Chile and the United States only.",
        f"Source workbook: `{SOURCE_WORKBOOK}`.",
        "Used sheets: " + ", ".join(f"`{sheet}`" for sheet in used_sheets) + ".",
        "",
        "## Detected structure",
        "",
        "Entity representation was detected from the wide data-sheet headers.",
        f"United States raw label: `{detected_entities.get('United States', 'NOT DETECTED')}`.",
        f"Chile raw label: `{detected_entities.get('Chile', 'NOT DETECTED')}`.",
        "Time variable: annual year in the first column.",
        f"Overall source year coverage: {format_year_range(*overall_year_range)}.",
        "",
        "## Output shape",
        "",
        "The processed CSV is wide.",
        "Rows are unique `entity`-`year` pairs.",
        "Columns after `standardized_entity`, `raw_entity_label`, and `year` preserve the original variable names exactly as they appear in the workbook’s title cells.",
        "Wide format is the conservative choice here because the source workbook itself stores one measure per sheet rather than a single long panel table.",
        "",
        "## Integrity checks",
        "",
        "Checks performed:",
        "- duplicate year rows inside each numeric sheet",
        "- duplicate entity-year-variable rows after normalization",
        "- duplicate entity-year rows in the final Chile–US panel",
        "- missingness by entity and variable after the 1900 filter",
        "- constant-series screening for the visual bundle",
        "- entity-label review against the auxiliary WBOP coverage sheet",
        "",
        (
            "Duplicate handling decision: no aggregation or collapse rule was applied because no duplicate conflicts required resolution."
            if duplicates_found.empty
            else "Duplicate handling decision: duplicate conflicts were detected and logged in the bundle outputs; they were not silently aggregated."
        ),
        "",
        "## Missing-data notes",
        "",
        "Missingness summary for Chile and the United States after 1900:",
        "",
        dataframe_to_markdown(
            ch_us_missing[
            [
                "source_sheet",
                "original_variable_name",
                "standardized_entity",
                "missingness_share",
                "min_year_available",
                "max_year_available",
            ]
            ]
        ),
        "",
        "## Variable dictionary summary",
        "",
        dataframe_to_markdown(
            variable_inventory_df[
            [
                "source_sheet",
                "original_variable_name",
                "included_in_visual_bundle",
                "included_in_chile_us_panel",
                "percent_missing",
            ]
            ]
        ),
        "",
        "## Caveats for later analytical use",
        "",
        "- Country units in the numeric sheets are coded (`US`, `GB`, `CL`) rather than fully spelled out.",
        "- The auxiliary `WBOP` sheet spells out names such as `USA`, `Britain`, and `Chile`; those naming differences were documented rather than forced into the numeric source.",
        "- Region labels vary slightly across some sheets. The Chile–US panel is unaffected because it uses only `US` and `CL`.",
        "- This file is an extraction and inspection product. It does not impose semantic harmonization beyond the explicit standardization of the entity labels in the processed output.",
    ]

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def make_plot(
    variable_df: pd.DataFrame,
    variable_name: str,
    source_sheet: str,
    png_path: Path,
    pdf_path: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(10, 6))

    years = variable_df["year"]
    min_year = int(years.min())
    max_year = int(years.max())

    for marker in HISTORICAL_MARKERS:
        if min_year <= marker <= max_year:
            ax.axvline(marker, color="0.85", linestyle="--", linewidth=0.8, zorder=0)

    for entity in ["United States", "United Kingdom", "Chile", "Latin America"]:
        entity_df = variable_df.loc[variable_df["standardized_entity"] == entity].sort_values("year")
        if entity_df["value"].notna().sum() == 0:
            continue
        ax.plot(entity_df["year"], entity_df["value"], label=entity, linewidth=1.6)

    ax.set_xlabel("Year")
    ax.set_ylabel("Value")
    ax.set_title(textwrap.fill(variable_name, width=72), fontsize=12)
    ax.text(
        0.0,
        1.02,
        f"Source sheet: {source_sheet}",
        transform=ax.transAxes,
        ha="left",
        va="bottom",
        fontsize=9,
        color="0.35",
    )
    ax.legend(frameon=False)
    ax.grid(True, color="0.92", linewidth=0.8)
    ax.set_axisbelow(True)
    fig.tight_layout()

    fig.savefig(png_path, dpi=200, bbox_inches="tight")
    fig.savefig(pdf_path, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    ensure_directories()
    logger = configure_logging(BUILD_LOG)

    logger.info("Starting WBOP inspection bundle build.")
    logger.info("Repo root: %s", REPO_ROOT)
    logger.info("Source workbook: %s", SOURCE_WORKBOOK)

    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE_WORKBOOK}")

    workbook = pd.ExcelFile(SOURCE_WORKBOOK)
    logger.info("Workbook sheets detected: %s", ", ".join(workbook.sheet_names))

    raw_sheets: dict[str, pd.DataFrame] = {}
    profiles: list[SheetProfile] = []
    normalized_sheets: dict[str, pd.DataFrame] = {}
    all_entity_labels: list[str] = []

    for sheet_name in workbook.sheet_names:
        raw_df = pd.read_excel(SOURCE_WORKBOOK, sheet_name=sheet_name, header=None)
        raw_sheets[sheet_name] = raw_df
        profile = inspect_sheet(sheet_name, raw_df)
        profiles.append(profile)
        logger.info(
            "Sheet profile | %s | role=%s | dims=%s | header=%s | time=%s | numeric_cols=%s",
            profile.sheet_name,
            profile.sheet_role,
            profile.dimensions,
            profile.header_structure,
            profile.likely_time_column,
            profile.numeric_columns,
        )
        if profile.sheet_role == "main_panel_source":
            normalized = normalize_data_sheet(raw_df, profile, sheet_name)
            normalized_sheets[sheet_name] = normalized
            all_entity_labels.extend([col for col in normalized.columns if col != "year"])

    used_sheets = list(normalized_sheets.keys())
    logger.info("Selected numeric data sheets: %s", ", ".join(used_sheets))

    wbop_text = " ".join(clean_string(x) for x in raw_sheets["WBOP"].fillna("").to_numpy().ravel())
    entity_mapping_df, detected_entities, manual_review_items = detect_target_entities(
        all_entity_labels,
        wbop_text,
        logger,
    )
    entity_mapping_df.to_csv(ENTITY_MAPPING_CSV, index=False)

    missing_targets = [entity for entity in TARGET_ENTITIES if entity not in detected_entities]
    if missing_targets:
        raise ValueError(f"Failed to detect required entities in workbook headers: {missing_targets}")

    filtered_long_frames: list[pd.DataFrame] = []
    duplicate_rows: list[dict[str, object]] = []
    excluded_rows: list[dict[str, object]] = []
    visual_inventory_rows: list[dict[str, object]] = []
    variable_inventory_rows: list[dict[str, object]] = []
    plotted_flags: dict[str, bool] = {}
    chile_us_frames: list[pd.DataFrame] = []

    overall_min_year = None
    overall_max_year = None

    for profile in profiles:
        if profile.sheet_role != "main_panel_source":
            continue

        sheet_name = profile.sheet_name
        wide_df = normalized_sheets[sheet_name]
        variable_name = profile.title
        time_years = wide_df["year"]
        if overall_min_year is None or int(time_years.min()) < overall_min_year:
            overall_min_year = int(time_years.min())
        if overall_max_year is None or int(time_years.max()) > overall_max_year:
            overall_max_year = int(time_years.max())

        sheet_duplicate_count = int(wide_df["year"].duplicated().sum())
        duplicate_rows.append(
            {
                "check_name": "duplicate_year_rows_sheet",
                "scope": "sheet",
                "source_sheet": sheet_name,
                "original_variable_name": variable_name,
                "duplicate_count": sheet_duplicate_count,
                "details": (
                    "No duplicated years detected in parsed sheet rows."
                    if sheet_duplicate_count == 0
                    else "Duplicated year rows detected in parsed sheet rows."
                ),
            }
        )

        selected_columns = ["year"] + list(detected_entities.values())
        subset_df = wide_df[selected_columns].copy()
        rename_map = {raw_label: std_label for std_label, raw_label in detected_entities.items()}
        subset_df = subset_df.rename(columns=rename_map)
        long_df = subset_df.melt(id_vars="year", var_name="standardized_entity", value_name="value")
        long_df["raw_entity_label"] = long_df["standardized_entity"].map(detected_entities)
        long_df["source_sheet"] = sheet_name
        long_df["original_variable_name"] = variable_name
        long_df = long_df[long_df["year"] >= 1900].sort_values(["standardized_entity", "year"]).reset_index(drop=True)
        filtered_long_frames.append(long_df)

        long_duplicate_count = int(
            long_df.duplicated(subset=["source_sheet", "standardized_entity", "year"]).sum()
        )
        duplicate_rows.append(
            {
                "check_name": "duplicate_entity_year_variable_rows_long",
                "scope": "long_filtered",
                "source_sheet": sheet_name,
                "original_variable_name": variable_name,
                "duplicate_count": long_duplicate_count,
                "details": (
                    "No duplicated entity-year-variable rows detected after normalization."
                    if long_duplicate_count == 0
                    else "Duplicated entity-year-variable rows detected after normalization."
                ),
            }
        )

        available_entities = sorted(
            long_df.loc[long_df["value"].notna(), "standardized_entity"].unique().tolist()
        )
        total_rows = len(long_df)
        nonmissing_rows = int(long_df["value"].notna().sum())
        missing_share = 1 - (nonmissing_rows / total_rows) if total_rows else None
        nonmissing_values = long_df.loc[long_df["value"].notna(), "value"]
        informative = False
        if not nonmissing_values.empty:
            entity_nunique = (
                long_df.dropna(subset=["value"])
                .groupby("standardized_entity")["value"]
                .nunique(dropna=True)
                .fillna(0)
            )
            informative = bool((entity_nunique > 1).any())

        exclusion_reason = None
        if nonmissing_rows == 0:
            exclusion_reason = "fully_missing_after_1900"
        elif not informative:
            exclusion_reason = "all_available_entity_series_constant_after_1900"

        plotted = exclusion_reason is None
        plotted_flags[sheet_name] = plotted

        if plotted:
            file_stem = f"{sheet_name}_{sanitize_filename(variable_name)}"
            png_relative = Path("figures") / "png" / f"{file_stem}.png"
            pdf_relative = Path("figures") / "pdf" / f"{file_stem}.pdf"
            png_path = OUTPUT_ROOT / png_relative
            pdf_path = OUTPUT_ROOT / pdf_relative
            try:
                make_plot(long_df, variable_name, sheet_name, png_path, pdf_path)
                visual_inventory_rows.append(
                    {
                        "original_variable_name": variable_name,
                        "sanitized_file_stem": file_stem,
                        "source_sheet": sheet_name,
                        "entities_available": ", ".join(available_entities),
                        "min_year": int(long_df.loc[long_df["value"].notna(), "year"].min()),
                        "max_year": int(long_df.loc[long_df["value"].notna(), "year"].max()),
                        "missingness_share": round(missing_share, 6) if missing_share is not None else None,
                        "png_path": png_relative.as_posix(),
                        "pdf_path": pdf_relative.as_posix(),
                        "plotted_yes_no": "yes",
                    }
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("Plotting failed for %s (%s)", sheet_name, variable_name)
                plotted_flags[sheet_name] = False
                excluded_rows.append(
                    {
                        "source_sheet": sheet_name,
                        "original_variable_name": variable_name,
                        "exclusion_reason": "plot_generation_failure",
                        "detail": str(exc),
                    }
                )
        else:
            excluded_rows.append(
                {
                    "source_sheet": sheet_name,
                    "original_variable_name": variable_name,
                    "exclusion_reason": exclusion_reason,
                    "detail": (
                        "No non-missing observations for the requested entities after 1900."
                        if exclusion_reason == "fully_missing_after_1900"
                        else "Available post-1900 observations were constant for every included entity."
                    ),
                }
            )

        chile_us_raw = {
            "United States": detected_entities["United States"],
            "Chile": detected_entities["Chile"],
        }
        chile_us_wide = wide_df[["year"] + list(chile_us_raw.values())].copy()
        chile_us_wide = chile_us_wide.rename(
            columns={
                chile_us_raw["United States"]: "United States",
                chile_us_raw["Chile"]: "Chile",
            }
        )
        chile_us_long = chile_us_wide.melt(id_vars="year", var_name="standardized_entity", value_name=variable_name)
        chile_us_long["raw_entity_label"] = chile_us_long["standardized_entity"].map(chile_us_raw)
        chile_us_frames.append(chile_us_long)

        panel_included = "yes"
        percent_missing = round(
            long_df.loc[long_df["standardized_entity"].isin(["Chile", "United States"]), "value"].isna().mean(),
            6,
        )
        nonmissing_years = long_df.loc[long_df["value"].notna(), "year"]
        variable_inventory_rows.append(
            {
                "original_variable_name": variable_name,
                "detected_type": "numeric",
                "role_guess": "time_series_measure",
                "source_sheet": sheet_name,
                "included_in_visual_bundle": "yes" if plotted_flags.get(sheet_name, False) else "no",
                "included_in_chile_us_panel": panel_included,
                "percent_missing": percent_missing,
                "min_year": int(nonmissing_years.min()) if not nonmissing_years.empty else None,
                "max_year": int(nonmissing_years.max()) if not nonmissing_years.empty else None,
                "notes": "Percent missing is calculated across Chile and United States after the 1900 filter.",
            }
        )

    filtered_long = pd.concat(filtered_long_frames, ignore_index=True)
    duplicate_df = pd.DataFrame(duplicate_rows).sort_values(
        ["check_name", "source_sheet"]
    ).reset_index(drop=True)
    duplicate_df.to_csv(DUPLICATE_CSV, index=False)

    missingness_df = build_missingness_summary(filtered_long, plotted_flags)
    missingness_df.to_csv(MISSINGNESS_CSV, index=False)

    excluded_df = pd.DataFrame(
        excluded_rows,
        columns=["source_sheet", "original_variable_name", "exclusion_reason", "detail"],
    )
    excluded_df.to_csv(EXCLUDED_SERIES_CSV, index=False)

    visual_inventory_df = pd.DataFrame(visual_inventory_rows).sort_values("source_sheet").reset_index(drop=True)
    visual_inventory_df.to_csv(VISUAL_INVENTORY_CSV, index=False)

    chile_us_long = pd.concat(chile_us_frames, ignore_index=True)
    chile_us_panel = (
        chile_us_long.pivot_table(
            index=["standardized_entity", "raw_entity_label", "year"],
            values=[row["original_variable_name"] for row in variable_inventory_rows],
            aggfunc="first",
        )
        .reset_index()
        .sort_values(["standardized_entity", "year"])
        .reset_index(drop=True)
    )
    panel_duplicate_count = int(
        chile_us_panel.duplicated(subset=["standardized_entity", "year"]).sum()
    )
    duplicate_df = pd.concat(
        [
            duplicate_df,
            pd.DataFrame(
                [
                    {
                        "check_name": "duplicate_entity_year_rows_processed_panel",
                        "scope": "processed_panel",
                        "source_sheet": "ALL",
                        "original_variable_name": "ALL",
                        "duplicate_count": panel_duplicate_count,
                        "details": (
                            "No duplicate entity-year rows detected in the processed Chile-US panel."
                            if panel_duplicate_count == 0
                            else "Duplicate entity-year rows detected in the processed Chile-US panel."
                        ),
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    duplicate_df.to_csv(DUPLICATE_CSV, index=False)
    chile_us_panel.to_csv(CHILE_US_PANEL_CSV, index=False)

    variable_inventory_df = pd.DataFrame(variable_inventory_rows).sort_values("source_sheet").reset_index(drop=True)
    inventory_summary_rows = [
        ["metric", "value"],
        ["source_workbook", str(SOURCE_WORKBOOK)],
        ["used_sheets", ", ".join(used_sheets)],
        ["visual_variables_plotted", int((variable_inventory_df["included_in_visual_bundle"] == "yes").sum())],
        ["variables_in_chile_us_panel", int((variable_inventory_df["included_in_chile_us_panel"] == "yes").sum())],
        ["post_1900_visual_observations", len(filtered_long)],
    ]
    write_variable_inventory_workbook(variable_inventory_df, inventory_summary_rows, VARIABLE_INVENTORY_XLSX)

    write_visual_markdown(
        VISUAL_MD,
        profiles,
        used_sheets,
        entity_mapping_df,
        detected_entities,
        visual_inventory_df,
        excluded_df,
        missingness_df,
        manual_review_items,
        duplicate_df,
        (overall_min_year, overall_max_year),
    )

    write_chile_us_user_guide(
        CHILE_US_GUIDE_MD,
        used_sheets,
        detected_entities,
        (overall_min_year, overall_max_year),
        duplicate_df,
        missingness_df,
        variable_inventory_df,
    )

    logger.info("Finished WBOP inspection bundle build.")
    logger.info("Inspection Markdown: %s", VISUAL_MD)
    logger.info("Chile-US panel: %s", CHILE_US_PANEL_CSV)
    logger.info("Variable inventory workbook: %s", VARIABLE_INVENTORY_XLSX)

    print("")
    print("WBOP inspection bundle complete")
    print(f"Workbook path: {SOURCE_WORKBOOK}")
    print(f"Sheets used: {', '.join(used_sheets)}")
    print("Detected entity column: wide sheet headers (country/regional labels across columns)")
    print("Detected time column: first column parsed as annual year")
    print(f"Number of observations after visual-bundle filtering: {len(filtered_long)}")
    print(f"Number of plotted variables: {int((visual_inventory_df['plotted_yes_no'] == 'yes').sum())}")
    print(f"Number of excluded variables: {len(excluded_df)}")
    print(f"Inspection bundle path: {VISUAL_MD}")
    print(f"Processed Chile-US panel path: {CHILE_US_PANEL_CSV}")
    print(f"Variable inventory workbook path: {VARIABLE_INVENTORY_XLSX}")


if __name__ == "__main__":
    main()
